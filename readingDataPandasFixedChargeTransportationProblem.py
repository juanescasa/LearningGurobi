# -*- coding: utf-8 -*-
"""
Created on Thu Apr 22 13:25:17 2021

@author: calle
"""
#This examples shows how to use pandas to read data from excel
#At the end it exports the results to an excel spreadsheet using openpyxl package
#it also imports a model define in a separate file to solve it

import time
import pandas as pd
import openpyxl
import fixedChargeTransportationModel

#I define this to compare how much it takes to read data with xlrd vs pandas
start_time_read_data = time.time()
#Read data
path="C:\OneDrive - Deakin University\OD\Personal\Software Documentation\Gurobi\data fixed charge transportation problem.xlsx"
sheet = 'shOrigins'
df_origins = pd.read_excel(path, sheet_name=sheet)

sOrigins=[]
sDestinations=[]
pSupply={}
pDemand={}
pCost={}
pOpenCost={}

for index in range(df_origins.shape[0]):
    i = df_origins['sOrigin'][index]
    sOrigins.append(i)
    pSupply[i]=df_origins['pSupply'][index]
    
sheet='shDestinations'
df_destinations = pd.read_excel(path, sheet_name=sheet)

for index in range(df_destinations.shape[0]):
    j = df_destinations['sDestination'][index]
    sDestinations.append(j)
    pDemand[j]=df_destinations['pDemand'][index]
    
sheet='shOriginsDestinations'
df_origins_destinations = pd.read_excel(path, sheet_name=sheet)

for index in range(df_origins_destinations.shape[0]):
    i = df_origins_destinations['sOrigin'][index]
    j = df_origins_destinations['sDestination'][index]
    pCost[i,j]=df_origins_destinations['pCost'][index]
    pOpenCost[i,j]=df_origins_destinations['pOpenCost'][index]

end_time_read_data = time.time()

print('read from pandas takes' + str(end_time_read_data-start_time_read_data))
print("--- %s seconds ---" % (end_time_read_data-start_time_read_data))

#The model function, apart from printing the solution it generates two outputs which are stored in a tuple
outputsModel = fixedChargeTransportationModel.transModel(sOrigins, sDestinations, pSupply, pDemand, pCost, pOpenCost)

#this calls the first output of the function that solves the model
outputVTransport = outputsModel[0]
outputVOpenArc = outputsModel[1]

#this code exports the excel spreadsheet with parameters and results in a sinle table
file="C:\OneDrive - Deakin University\OD\Personal\Software Documentation\Gurobi\output fixed charge transportation problem v2.xlsx"
sheet="shOutput"
workbook = openpyxl.load_workbook(file)
sheet_to_remove=workbook[sheet]
workbook.remove(sheet_to_remove)
ws = workbook.create_sheet(sheet)
ws.cell(row=1, column=1, value="Origin")
ws.cell(row=1, column=2, value="Destination")
ws.cell(row=1, column=3, value="Transportation Cost")
ws.cell(row=1, column=4, value="Arc Opening Cost")
ws.cell(row=1, column=5, value="Transport")
ws.cell(row=1, column=6, value="Open Arc")

index_row=2
for i in sOrigins:
    for j in sDestinations:
        ws.cell(row=index_row, column=1, value=i)
        ws.cell(row=index_row, column=2, value=j)
        ws.cell(row=index_row, column=3, value=pCost[i,j])
        ws.cell(row=index_row, column=4, value=pOpenCost[i,j])
        ws.cell(row=index_row, column=5, value=outputVTransport[i,j])
        ws.cell(row=index_row, column=6, value=outputVOpenArc[i,j])
        index_row=index_row+1
       
workbook.save(file)          

